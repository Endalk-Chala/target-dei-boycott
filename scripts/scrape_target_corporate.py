#!/usr/bin/env python3
"""
Target DEI Boycott — Corporate Communications Data Collector & Scraper

This script:
1. Compiles all known Target corporate communications related to the
   DEI rollback and boycott from Jan 2025 through Apr 2026
2. Scrapes corporate.target.com press releases for additional items
3. Outputs a structured, coding-ready Excel spreadsheet

Author: Endalkachew H. Chala, PhD
Date: April 2026

Requirements: pip install pandas openpyxl requests beautifulsoup4
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    import requests
    from bs4 import BeautifulSoup
    HAS_SCRAPING = True
except ImportError:
    HAS_SCRAPING = False
    print("NOTE: requests/beautifulsoup4 not installed. Skipping live scraping.")
    print("      Install with: pip install requests beautifulsoup4")

# ══════════════════════════════════════════════════════════════════
# COMMUNICATION TYPE CODES (matching codebook)
# ══════════════════════════════════════════════════════════════════
COMM_TYPES = {
    1: "Press Release/Fact Sheet",
    2: "Earnings Call/Investor",
    3: "Social Media Post",
    4: "Executive Statement/Interview",
    5: "Website Content Change",
    6: "Media Response/Statement",
    7: "Internal Memo (leaked/reported)",
}

SPOKESPERSON_CODES = {
    1: "CEO Brian Cornell",
    2: "Kiera Fernandez (Chief Community Impact & Equity Officer)",
    3: "CEO Michael Fiddelke",
    4: "Corporate Spokesperson (unnamed)",
    5: "Social Media Team",
    6: "Board of Directors",
    7: "Other Executive",
}

# ══════════════════════════════════════════════════════════════════
# PRE-POPULATED CORPORATE COMMUNICATIONS DATABASE
# Built from web research of Target corporate site and news coverage
# ══════════════════════════════════════════════════════════════════

communications = [
    {
        "Unit_ID": 1,
        "Date": "2025-01-24",
        "Comm_Type": 1,
        "Comm_Type_Label": "Press Release/Fact Sheet",
        "Title": "Target's Belonging at the Bullseye Strategy",
        "Spokesperson": 2,
        "Spokesperson_Label": "Kiera Fernandez",
        "Source_URL": "https://corporate.target.com/press/fact-sheet/2025/01/belonging-bullseye-strategy",
        "Source_Type": "Corporate Website",
        "Full_Text_Excerpt": (
            "We remain focused on driving our business by creating a sense of belonging "
            "for our team, guests and communities through a commitment to inclusion. "
            "Throughout 2025, we'll be accelerating action in key areas and implementing "
            "changes with the goal of driving growth and staying in step with the evolving "
            "external landscape."
        ),
        "Key_Actions": (
            "Concluding three-year DEI goals; Concluding REACH initiatives; "
            "Stopping external diversity-focused surveys including HRC CEI; "
            "Evolving Supplier Diversity to Supplier Engagement; "
            "Refocusing employee resource groups on development and mentorship"
        ),
        "Key_Language": (
            "belonging; staying in step with the evolving external landscape; "
            "inclusion; driving growth; consumer relevance; business results"
        ),
        "Boycott_Period": "Pre-Boycott (Trigger Event)",
        "Notes": (
            "CEO Brian Cornell notably ABSENT from this announcement. "
            "Kiera Fernandez delivered the news. This is the trigger event for the boycott. "
            "Announced just 4 days into Trump's second presidency."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 2,
        "Date": "2025-02-01",
        "Comm_Type": 5,
        "Comm_Type_Label": "Website Content Change",
        "Title": "Target corporate website — Belonging page update",
        "Spokesperson": 4,
        "Spokesperson_Label": "Corporate (website)",
        "Source_URL": "https://corporate.target.com/sustainability-governance/our-team/belonging",
        "Source_Type": "Corporate Website",
        "Full_Text_Excerpt": (
            "Page rebranded from Diversity, Equity and Inclusion to Belonging at the Bullseye. "
            "Language shifted from DEI commitments to belonging and inclusion framing."
        ),
        "Key_Actions": "Website restructure; DEI language removal; Belonging rebrand",
        "Key_Language": "belonging; inclusion; our team; our guests; our communities",
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "Check Wayback Machine for exact date of change. The DEI page at "
            "corporate.target.com/sustainability-governance/our-team/diversity-equity-inclusion "
            "now redirects to the Belonging page."
        ),
        "Verified": False,
    },
    {
        "Unit_ID": 3,
        "Date": "2025-02-04",
        "Comm_Type": 3,
        "Comm_Type_Label": "Social Media Post",
        "Title": "Target Black History Month social media content",
        "Spokesperson": 5,
        "Spokesperson_Label": "Social Media Team",
        "Source_URL": "",
        "Source_Type": "Social Media (Instagram/TikTok/Facebook)",
        "Full_Text_Excerpt": "",
        "Key_Actions": "Black History Month promotional content posted during active boycott",
        "Key_Language": "",
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "Context: Boycott launched during Black History Month. "
            "Check Target Instagram/TikTok/Facebook for BHM-related content. "
            "Top comments dominated by boycott supporters per news reports."
        ),
        "Verified": False,
    },
    {
        "Unit_ID": 4,
        "Date": "2025-02-25",
        "Comm_Type": 6,
        "Comm_Type_Label": "Media Response/Statement",
        "Title": "Target statement on continued Black-owned product offerings",
        "Spokesperson": 4,
        "Spokesperson_Label": "Corporate Spokesperson",
        "Source_URL": "",
        "Source_Type": "News Coverage (CNN, multiple outlets)",
        "Full_Text_Excerpt": (
            "Target said in February that it continues to offer products from "
            "Black-owned and minority-owned businesses, adding that its Black History "
            "Month offerings were available in stores and online."
        ),
        "Key_Actions": "Affirmed continued Black-owned product availability",
        "Key_Language": "continues to offer; Black History Month offerings; available",
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "Reported by CNN. Target did not directly address the boycott in this statement. "
            "Statement focused on product availability rather than policy."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 5,
        "Date": "2025-03-05",
        "Comm_Type": 2,
        "Comm_Type_Label": "Earnings Call/Investor",
        "Title": "Q4 FY2024 Earnings Call",
        "Spokesperson": 1,
        "Spokesperson_Label": "CEO Brian Cornell",
        "Source_URL": "",
        "Source_Type": "Earnings Call Transcript",
        "Full_Text_Excerpt": "",
        "Key_Actions": "Earnings call during active boycott period",
        "Key_Language": "",
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "Retail Brew reported Target 'ignores DEI backlash, boycott' in this call. "
            "NEED: Full transcript to code DEI/boycott references. "
            "Check if Cornell or any exec addressed boycott during Q&A. "
            "Source: retailbrew.com/stories/2025/03/05/target-earnings-call-ignores-dei-backlash-boycott"
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 6,
        "Date": "2025-04-17",
        "Comm_Type": 4,
        "Comm_Type_Label": "Executive Statement/Interview",
        "Title": "Brian Cornell meets with Rev. Jamal Bryant — $2B commitment affirmed",
        "Spokesperson": 1,
        "Spokesperson_Label": "CEO Brian Cornell",
        "Source_URL": "",
        "Source_Type": "News Coverage (CNN, Apr 21 2025)",
        "Full_Text_Excerpt": (
            "After meeting with Target CEO Brian Cornell on Thursday, Bryant announced "
            "Sunday that Target agreed to honor its pledge to spend $2 billion with "
            "Black-owned businesses — a commitment initially made in 2021."
        ),
        "Key_Actions": "CEO met with boycott leader; Affirmed $2B Black-owned business commitment",
        "Key_Language": "honor its pledge; $2 billion; Black-owned businesses",
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "Reported by CNN Apr 21. Cornell met Bryant on Apr 17 (Thursday before Easter). "
            "Target did not respond to CNN's latest request for comment after this meeting. "
            "Bryant called for boycott to continue despite this concession."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 7,
        "Date": "2025-04-21",
        "Comm_Type": 6,
        "Comm_Type_Label": "Media Response/Statement",
        "Title": "Target non-response to CNN on boycott impact",
        "Spokesperson": 4,
        "Spokesperson_Label": "Corporate Spokesperson",
        "Source_URL": "https://www.cnn.com/2025/04/21/business/target-boycott-jamal-bryant-minority-businesses",
        "Source_Type": "News Coverage (CNN)",
        "Full_Text_Excerpt": "Target did not respond to CNN's latest request for comment.",
        "Key_Actions": "No comment on boycott; Foot traffic down 9% Feb, 6.5% Mar (Placer.ai)",
        "Key_Language": "silence; non-response",
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "Significant: Target chose silence rather than public response to boycott coverage. "
            "CNN reported 10 consecutive weeks of foot traffic decline. "
            "Placer.ai data: -9% YoY February, -6.5% YoY March."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 8,
        "Date": "2025-05-21",
        "Comm_Type": 2,
        "Comm_Type_Label": "Earnings Call/Investor",
        "Title": "Q1 FY2025 Earnings Call — First acknowledgment of DEI impact",
        "Spokesperson": 1,
        "Spokesperson_Label": "CEO Brian Cornell",
        "Source_URL": "",
        "Source_Type": "Earnings Call Transcript / News Coverage (CNN, MSNBC)",
        "Full_Text_Excerpt": (
            "Cornell said the company 'faced several additional headwinds this quarter, "
            "including five consecutive months of declining consumer confidence, uncertainty "
            "regarding the impact of potential tariffs, and the reaction to the updates "
            "we shared on [DEI] in January.'"
        ),
        "Key_Actions": (
            "First explicit acknowledgment of DEI backlash impact on business; "
            "Comp sales down 3.8%"
        ),
        "Key_Language": (
            "headwinds; reaction to the updates we shared on [DEI]; "
            "declining consumer confidence; tariffs"
        ),
        "Boycott_Period": "Peak Boycott (Jan-Apr 2025)",
        "Notes": (
            "CRITICAL COMMUNICATION: First time Target explicitly linked DEI rollback to "
            "business performance. Framed as one of several 'headwinds' alongside tariffs "
            "and consumer confidence. CNN reported 'sales fell last quarter, driven in part "
            "by customer backlash.' Sales down nearly 3% in Q1."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 9,
        "Date": "2025-08-20",
        "Comm_Type": 1,
        "Comm_Type_Label": "Press Release",
        "Title": "Target Appoints Michael Fiddelke As Chief Executive Officer",
        "Spokesperson": 6,
        "Spokesperson_Label": "Board of Directors",
        "Source_URL": "https://corporate.target.com/press/release/2025/08/target-appoints-michael-fiddelke-as-chief-executive-officer",
        "Source_Type": "Corporate Website",
        "Full_Text_Excerpt": (
            "Target Corporation today announced the company's Board of Directors has "
            "unanimously elected Michael Fiddelke, chief operating officer, to succeed "
            "Brian Cornell as chief executive officer. Both appointments effective Feb. 1, 2026."
        ),
        "Key_Actions": (
            "CEO succession announced; Cornell to Executive Chair; "
            "Fiddelke named next CEO effective Feb 1 2026"
        ),
        "Key_Language": (
            "new chapter; return Target to growth; refocus and accelerate; "
            "reestablish Target's position; fresh eyes mindset; challenging the status quo"
        ),
        "Boycott_Period": "Mid-Boycott (May-Dec 2025)",
        "Notes": (
            "No mention of DEI or boycott in the announcement. "
            "Fortune reported boycott organizers said 'leadership change doesn't mean "
            "anything without a culture change.' Fiddelke later met with boycott leaders."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 10,
        "Date": "2025-10-24",
        "Comm_Type": 6,
        "Comm_Type_Label": "Media Response/Statement",
        "Title": "Target spotlights support for Black founders — Fortune report",
        "Spokesperson": 4,
        "Spokesperson_Label": "Corporate Spokesperson",
        "Source_URL": "",
        "Source_Type": "News Coverage (Fortune)",
        "Full_Text_Excerpt": "",
        "Key_Actions": "Target highlighted Black-founded brands after sustained DEI backlash",
        "Key_Language": "Black founders; support",
        "Boycott_Period": "Mid-Boycott (May-Dec 2025)",
        "Notes": (
            "Fortune reported Target spotlighting Black-founded brands. "
            "Potential shift toward conciliatory/restorative stance. "
            "Compare framing to January 2025 announcement. Verify exact date and content."
        ),
        "Verified": False,
    },
    {
        "Unit_ID": 11,
        "Date": "2025-11-19",
        "Comm_Type": 2,
        "Comm_Type_Label": "Earnings Call/Investor",
        "Title": "Q3 FY2025 Earnings Call",
        "Spokesperson": 1,
        "Spokesperson_Label": "CEO Brian Cornell",
        "Source_URL": "https://corporate.target.com/press/release/2025/11/target-corporation-reports-third-quarter-earnings",
        "Source_Type": "Corporate Website / Earnings Call Transcript",
        "Full_Text_Excerpt": "",
        "Key_Actions": "Quarterly earnings during ongoing boycott",
        "Key_Language": "",
        "Boycott_Period": "Mid-Boycott (May-Dec 2025)",
        "Notes": (
            "NEED: Full transcript to code any DEI/boycott references. "
            "Check if boycott acknowledged or if layoffs (1,800 corporate) discussed "
            "in context of reputational damage."
        ),
        "Verified": False,
    },
    {
        "Unit_ID": 12,
        "Date": "2026-02-01",
        "Comm_Type": 4,
        "Comm_Type_Label": "Executive Statement",
        "Title": "Michael Fiddelke takes over as CEO",
        "Spokesperson": 3,
        "Spokesperson_Label": "CEO Michael Fiddelke",
        "Source_URL": "",
        "Source_Type": "Corporate / News Coverage",
        "Full_Text_Excerpt": "",
        "Key_Actions": "CEO transition from Cornell to Fiddelke",
        "Key_Language": "",
        "Boycott_Period": "Resolution (Jan-Apr 2026)",
        "Notes": (
            "Fiddelke met with boycott leaders in recent months before taking over. "
            "Axios described him as 'very keen to listen and to learn from all kinds "
            "of stakeholders.' Check for any public statements on DEI/belonging."
        ),
        "Verified": False,
    },
    {
        "Unit_ID": 13,
        "Date": "2026-03-03",
        "Comm_Type": 1,
        "Comm_Type_Label": "Press Release/Strategic Plan",
        "Title": "Target Outlines Strategic Plan for New Chapter of Growth in 2026 and Beyond",
        "Spokesperson": 3,
        "Spokesperson_Label": "CEO Michael Fiddelke",
        "Source_URL": "https://corporate.target.com/news-features/article/2026/03/target-growth-strategy-2026",
        "Source_Type": "Corporate Website",
        "Full_Text_Excerpt": (
            "Four clear priorities: Lead with merchandising authority; "
            "Accelerate technology; Elevate the guest experience; "
            "Strengthen team and communities. $2 billion incremental investment in 2026."
        ),
        "Key_Actions": (
            "$2B incremental investment; 30+ new stores; 130+ remodels; "
            "'Strengthen team and communities' as strategic priority"
        ),
        "Key_Language": (
            "new chapter; growth; strengthen team and communities; "
            "long-standing commitment to communities; belonging NOT mentioned"
        ),
        "Boycott_Period": "Resolution (Jan-Apr 2026)",
        "Notes": (
            "Notable: 'Strengthen team and communities' is one of 4 strategic pillars "
            "but NO explicit mention of DEI, diversity, equity, inclusion, or belonging. "
            "Compare language carefully to Jan 2025 'Belonging at the Bullseye.' "
            "Released just 8 days before boycott officially ended."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 14,
        "Date": "2026-03-11",
        "Comm_Type": 6,
        "Comm_Type_Label": "Media Response/Statement",
        "Title": "Target statement on boycott ending",
        "Spokesperson": 4,
        "Spokesperson_Label": "Corporate Spokesperson",
        "Source_URL": "https://www.axios.com/2026/03/11/target-boycott-ends-dei",
        "Source_Type": "News Coverage (Axios)",
        "Full_Text_Excerpt": (
            "'Target is more committed than ever to creating growth and opportunity "
            "for all,' the retailer said in a statement to Axios, also noting it "
            "serves more than 2,000 communities."
        ),
        "Key_Actions": (
            "Official statement on boycott ending; Stated no policies reversed or "
            "reinstated; Described $2B pledge as completion of existing commitment"
        ),
        "Key_Language": (
            "growth and opportunity for all; more committed than ever; "
            "2,000 communities; no policies reversed or reinstated"
        ),
        "Boycott_Period": "Resolution (Jan-Apr 2026)",
        "Notes": (
            "CRITICAL: Target explicitly stated 'no policies were reversed or reinstated "
            "as a result of conversations with its leadership.' Framed $2B pledge as "
            "'completion of an existing commitment' not a concession. Boycott leaders "
            "Tamika Mallory called for public apology that 'has not happened.' "
            "Minnesota local groups said local boycott still ongoing."
        ),
        "Verified": True,
    },
    {
        "Unit_ID": 15,
        "Date": "2026-03-26",
        "Comm_Type": 6,
        "Comm_Type_Label": "Media Response/Statement",
        "Title": "Target faces new boycott over ICE response",
        "Spokesperson": 4,
        "Spokesperson_Label": "Corporate Spokesperson",
        "Source_URL": "https://www.cnbc.com/2026/03/26/target-aft-boycott-ice-minneapolis.html",
        "Source_Type": "News Coverage (CNBC)",
        "Full_Text_Excerpt": "",
        "Key_Actions": "New boycott threat from AFT over ICE-related response in Minneapolis",
        "Key_Language": "",
        "Boycott_Period": "Resolution (Jan-Apr 2026)",
        "Notes": (
            "Post-DEI boycott, new controversy emerged. CNBC reported Target faced "
            "another boycott threat related to ICE response. Shows ongoing reputational "
            "vulnerability. Minneapolis connection relevant to Minnesota angle."
        ),
        "Verified": False,
    },
]


def build_dataset():
    """Build the corporate communications dataset."""
    print("=" * 60)
    print("TARGET DEI BOYCOTT — CORPORATE COMMUNICATIONS COLLECTOR")
    print("=" * 60)
    print(f"\nPre-populated entries: {len(communications)}")

    df = pd.DataFrame(communications)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").reset_index(drop=True)
    df["Unit_ID"] = range(1, len(df) + 1)

    return df


def scrape_target_press(df):
    """Attempt to scrape additional press releases from corporate.target.com."""
    if not HAS_SCRAPING:
        print("\nSkipping live scraping (missing dependencies).")
        return df

    print("\nAttempting to scrape corporate.target.com press releases...")

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }

    # Target press release listing pages to check
    urls_to_check = [
        "https://corporate.target.com/press/releases",
        "https://corporate.target.com/news-features",
    ]

    new_items = []
    existing_urls = set(df["Source_URL"].dropna().values)

    for url in urls_to_check:
        try:
            print(f"  Checking: {url}")
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code != 200:
                print(f"    Status {resp.status_code}, skipping.")
                continue

            soup = BeautifulSoup(resp.text, "html.parser")

            # Find links that might be press releases
            links = soup.find_all("a", href=True)
            for link in links:
                href = link.get("href", "")
                text = link.get_text(strip=True)

                # Filter for DEI/diversity/belonging related content
                keywords = [
                    "dei", "diversity", "equity", "inclusion", "belonging",
                    "boycott", "community", "supplier", "black-owned",
                    "racial", "reach", "bullseye",
                ]
                text_lower = text.lower()
                href_lower = href.lower()

                if any(kw in text_lower or kw in href_lower for kw in keywords):
                    full_url = href if href.startswith("http") else f"https://corporate.target.com{href}"

                    if full_url not in existing_urls:
                        print(f"    FOUND: {text[:80]}...")
                        print(f"           {full_url}")
                        new_items.append({
                            "Title": text[:200],
                            "Source_URL": full_url,
                            "Source_Type": "Corporate Website (scraped)",
                            "Notes": "Auto-discovered — needs manual review and coding",
                            "Verified": False,
                        })
                        existing_urls.add(full_url)

        except Exception as e:
            print(f"    Error: {e}")

    if new_items:
        print(f"\n  Found {len(new_items)} potential new items to review.")
    else:
        print("\n  No new items found via scraping.")

    return new_items


def create_coding_spreadsheet(df, new_items, output_path):
    """Create a professionally formatted coding-ready Excel spreadsheet."""
    print(f"\nCreating coding spreadsheet: {output_path}")

    wb = Workbook()

    # ── Sheet 1: Corporate Communications (main coding sheet) ──
    ws = wb.active
    ws.title = "Corporate_Comms"

    # Header styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Column definitions
    columns = [
        ("Unit_ID", 8),
        ("Date", 12),
        ("Comm_Type", 10),
        ("Comm_Type_Label", 25),
        ("Title", 45),
        ("Spokesperson", 12),
        ("Spokesperson_Label", 30),
        ("Source_URL", 40),
        ("Source_Type", 20),
        ("Full_Text_Excerpt", 50),
        ("Key_Actions", 40),
        ("Key_Language", 35),
        ("Boycott_Period", 22),
        ("DEI_Framing", 12),
        ("Boycott_Ack", 12),
        ("Community_Stance", 14),
        ("Tone", 12),
        ("Notes", 50),
        ("Verified", 10),
        ("Coder_Name", 15),
        ("Coder_Date", 12),
    ]

    # Write headers
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(vertical="top", wrap_text=True)
    verified_fill = PatternFill("solid", fgColor="E2EFDA")
    unverified_fill = PatternFill("solid", fgColor="FFF2CC")

    for row_idx, record in df.iterrows():
        row_num = row_idx + 2
        for col_idx, (col_name, _) in enumerate(columns, 1):
            value = record.get(col_name, "")
            if col_name == "Date" and pd.notna(value):
                value = pd.Timestamp(value).strftime("%Y-%m-%d")
            elif pd.isna(value):
                value = ""

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

            # Color rows by verification status
            if record.get("Verified", False):
                cell.fill = verified_fill
            else:
                cell.fill = unverified_fill

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # ── Sheet 2: Scraped Items (to review) ──
    ws2 = wb.create_sheet("Scraped_Items_Review")
    scraped_cols = [("Title", 50), ("Source_URL", 60), ("Source_Type", 25), ("Notes", 50)]

    for col_idx, (col_name, width) in enumerate(scraped_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = header_align
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, item in enumerate(new_items, 2):
        for col_idx, (col_name, _) in enumerate(scraped_cols, 1):
            ws2.cell(row=row_idx, column=col_idx, value=item.get(col_name, ""))

    # ── Sheet 3: Coding Reference ──
    ws3 = wb.create_sheet("Coding_Reference")

    # Communication types
    ws3.cell(row=1, column=1, value="Communication Type Codes").font = Font(bold=True, size=11)
    ws3.cell(row=2, column=1, value="Code").font = Font(bold=True)
    ws3.cell(row=2, column=2, value="Label").font = Font(bold=True)
    for i, (code, label) in enumerate(COMM_TYPES.items(), 3):
        ws3.cell(row=i, column=1, value=code)
        ws3.cell(row=i, column=2, value=label)

    row_start = i + 2
    ws3.cell(row=row_start, column=1, value="Spokesperson Codes").font = Font(bold=True, size=11)
    ws3.cell(row=row_start + 1, column=1, value="Code").font = Font(bold=True)
    ws3.cell(row=row_start + 1, column=2, value="Label").font = Font(bold=True)
    for i, (code, label) in enumerate(SPOKESPERSON_CODES.items(), row_start + 2):
        ws3.cell(row=i, column=1, value=code)
        ws3.cell(row=i, column=2, value=label)

    row_start = i + 2
    ws3.cell(row=row_start, column=1, value="Coding Variables (to be filled)").font = Font(bold=True, size=11)
    coding_vars = [
        ("DEI_Framing", "1=Proactive/Positive, 2=Defensive, 3=Dismissive, 4=Absent/Silent, 5=Reframing, 99=Cannot Determine"),
        ("Boycott_Ack", "0=Not Acknowledged, 1=Implicit Reference, 2=Explicit Acknowledgment, 3=Detailed Response"),
        ("Community_Stance", "0=No Engagement, 1=Dismissive, 2=Neutral/Procedural, 3=Conciliatory, 4=Restorative"),
        ("Tone", "1=Corporate/Formal, 2=Defensive, 3=Conciliatory, 4=Forward-looking, 5=Empathetic, 99=Cannot Determine"),
    ]
    for i, (var, desc) in enumerate(coding_vars, row_start + 1):
        ws3.cell(row=i, column=1, value=var).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=desc)

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 80

    # ── Sheet 4: Timeline ──
    ws4 = wb.create_sheet("Timeline")
    ws4.cell(row=1, column=1, value="Date").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="Event").font = Font(bold=True)
    ws4.cell(row=1, column=3, value="Period").font = Font(bold=True)
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 70
    ws4.column_dimensions["C"].width = 25

    timeline_events = [
        ("2025-01-20", "Trump inaugurated for second term", "Context"),
        ("2025-01-24", "Target announces 'Belonging at the Bullseye' — ends DEI goals, REACH, HRC CEI", "Trigger"),
        ("2025-02-04", "Black History Month begins during active backlash", "Peak Boycott"),
        ("2025-03-05", "40-day Lent 'Target Fast' boycott begins (Rev. Jamal Bryant)", "Peak Boycott"),
        ("2025-03-05", "Q4 FY2024 Earnings — boycott not addressed", "Peak Boycott"),
        ("2025-04-17", "Cornell meets Bryant — affirms $2B Black-owned commitment", "Peak Boycott"),
        ("2025-04-17", "Target Fast 40-day period ends", "Peak Boycott"),
        ("2025-04-21", "CNN reports 10 weeks of declining foot traffic; Target silent", "Peak Boycott"),
        ("2025-05-21", "Q1 FY2025 Earnings — Cornell first acknowledges DEI impact; sales down 3.8%", "Peak Boycott"),
        ("2025-08-20", "Fiddelke named next CEO; Cornell to Executive Chair", "Mid-Boycott"),
        ("2025-10-24", "Target spotlights Black founders (Fortune report)", "Mid-Boycott"),
        ("2025-11-19", "Q3 FY2025 Earnings", "Mid-Boycott"),
        ("2025-12-31", "~1,800 corporate layoffs announced", "Mid-Boycott"),
        ("2026-02-01", "Fiddelke officially becomes CEO", "Resolution"),
        ("2026-03-03", "Target announces 2026 growth strategy — $2B investment", "Resolution"),
        ("2026-03-11", "Boycott leaders announce campaign 'officially concluding'", "Resolution"),
        ("2026-03-11", "Target states 'no policies reversed or reinstated'", "Resolution"),
        ("2026-03-13", "Bryant retracts statement; MN local groups say boycott continues", "Resolution"),
        ("2026-03-26", "New AFT boycott threat over ICE response in Minneapolis", "Post-Resolution"),
    ]

    for i, (date, event, period) in enumerate(timeline_events, 2):
        ws4.cell(row=i, column=1, value=date)
        ws4.cell(row=i, column=2, value=event)
        ws4.cell(row=i, column=3, value=period)

    # Save
    wb.save(output_path)
    print(f"  Saved with {len(df)} communications across 4 sheets.")


def main():
    """Run the full collection pipeline."""
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "Target_Corporate_Comms_Dataset.xlsx")

    # Build pre-populated dataset
    df = build_dataset()

    # Attempt live scraping
    new_items = scrape_target_press(df) if HAS_SCRAPING else []

    # Create coding spreadsheet
    create_coding_spreadsheet(df, new_items if isinstance(new_items, list) else [], output_path)

    # Summary
    verified = df["Verified"].sum()
    unverified = len(df) - verified
    print(f"\n{'=' * 60}")
    print("COLLECTION SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Total communications: {len(df)}")
    print(f"  Verified (green):     {verified}")
    print(f"  Need verification:    {unverified} (yellow)")
    print(f"  Scraped items review: {len(new_items) if isinstance(new_items, list) else 0}")
    print(f"\n  Output: {output_path}")
    print(f"\nNEXT STEPS:")
    print(f"  1. Open the spreadsheet and verify yellow-highlighted entries")
    print(f"  2. Add Source_URLs for entries missing them")
    print(f"  3. Fill in Full_Text_Excerpt for entries that need it")
    print(f"  4. Code DEI_Framing, Boycott_Ack, Community_Stance, Tone for all entries")
    print(f"  5. Review Scraped_Items_Review sheet for any additional communications")
    print(f"  6. Save coded version as 'Target_Corporate_Comms_Coded.xlsx'")


if __name__ == "__main__":
    main()
